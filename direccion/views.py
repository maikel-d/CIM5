# ============================================================
# Sistema de Gestión - Dirección General
# views.py
# ============================================================

# ============================================================
# Python Standard Library
# ============================================================
import json
import zipfile
import shutil
import tempfile
from io import BytesIO
from pathlib import Path
from datetime import datetime, timedelta, date
import calendar
import re

# ============================================================
# Django Core
# ============================================================
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Case, When, IntegerField, Count
from django.db.models.functions import TruncMonth
from django.conf import settings
from django.urls import reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.middleware.csrf import get_token

# ============================================================
# Third-party
# ============================================================
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.colors import HexColor
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image
)

# ============================================================
# Local App
# ============================================================
from .models import (
    UserProfile, Personal, DocumentoPersonal,
    Caso, Investigado, DocumentoInvestigado, DocumentoDireccion, DocumentoCaso, AuditLog,
    Tarea, TicketSoporte, Notificacion, notificar_administradores, InformeDiario,
    Bien, DocumentoBien
)
from .forms import (
    LoginForm, UserCreateForm, UserEditForm,
    PersonalForm, DocumentoPersonalForm,
    CasoForm, InvestigadoForm, DocumentoInvestigadoForm,
    DocumentoDireccionForm, DocumentoCasoForm,
    TareaForm, TicketSoporteForm, TicketAsignarForm,
    InformeDiarioForm, BienForm, DocumentoBienForm
)
from .decorators import permiso_required, _pagina_403
from .audit import auditar
from . import permissions as perms


# ============================================================
# Mixins
# ============================================================

class PermissionRequiredMixin(LoginRequiredMixin):
    """Mixin que verifica permisos granulares en dispatch().
    Definir permisos_requeridos con la lista de permisos necesarios.
    El usuario necesita tener AL MENOS UNO de ellos para acceder.
    Hereda de LoginRequiredMixin para manejar usuarios no autenticados.
    """
    permisos_requeridos = []

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        if request.user.is_superuser:
            return super().dispatch(request, *args, **kwargs)
        try:
            profile = request.user.profile
            for permiso in self.permisos_requeridos:
                if profile.tiene_permiso(permiso):
                    return super().dispatch(request, *args, **kwargs)
            return _pagina_403(request.user, "No tienes permisos para acceder a esta sección.", get_token(request))
        except Exception:
            return _pagina_403(request.user, "No tienes permisos para acceder a esta sección.", get_token(request))


# ============================================================
# AUTHENTICATION
# ============================================================

class CustomLoginView(View):
    template_name = "login.html"

    def get(self, request):
        if request.user.is_authenticated:
            return redirect("dashboard")
        form = LoginForm()
        return render(request, self.template_name, {"form": form})

    def post(self, request):
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data["username"]
            password = form.cleaned_data["password"]
            user = authenticate(request, username=username, password=password)
            if user is not None:
                if user.is_active:
                    login(request, user)
                    messages.success(request, f"Bienvenido, {user.get_full_name() or user.username}")
                    return redirect("dashboard")
                else:
                    messages.error(request, "Esta cuenta está desactivada. Contacte al administrador.")
            else:
                messages.error(request, "Usuario o contraseña incorrectos.")
        return render(request, self.template_name, {"form": form})


# ============================================================
# DASHBOARD
# ============================================================

@permiso_required(perms.DASHBOARD_VER)
def dashboard(request):
    total_personal = Personal.objects.filter(activo=True).count()
    total_casos = Caso.objects.filter(activo=True).count()
    total_investigados = Investigado.objects.filter(activo=True).count()
    total_usuarios = User.objects.filter(is_active=True).count()
    total_bienes = Bien.objects.filter(activo=True).count()
    tareas_pendientes = Tarea.objects.filter(completada=False).order_by("-fecha_creacion")
    tareas_count = Tarea.objects.count()
    tareas_completadas = Tarea.objects.filter(completada=True).count()

    # Document type counts — 4 queries using conditional aggregation (instead of 12+)
    _doc_q_pdf = Count('id', filter=Q(tipo='PDF'))
    _doc_q_word = Count('id', filter=Q(tipo='WORD'))
    _doc_q_img = Count('id', filter=Q(tipo='IMAGEN'))
    _doc_q_otro = Count('id', filter=Q(tipo='OTRO'))
    dp = DocumentoPersonal.objects.aggregate(p=_doc_q_pdf, w=_doc_q_word, i=_doc_q_img, o=_doc_q_otro)
    di = DocumentoInvestigado.objects.aggregate(p=_doc_q_pdf, w=_doc_q_word, i=_doc_q_img, o=_doc_q_otro)
    dc = DocumentoCaso.objects.aggregate(p=_doc_q_pdf, w=_doc_q_word, i=_doc_q_img, o=_doc_q_otro)
    dd = DocumentoDireccion.objects.aggregate(p=_doc_q_pdf, w=_doc_q_word, i=_doc_q_img, o=_doc_q_otro)
    docs_pdf = dp['p'] + di['p'] + dc['p'] + dd['p']
    docs_word = dp['w'] + di['w'] + dc['w'] + dd['w']
    docs_img = dp['i'] + di['i'] + dc['i'] + dd['i']
    total_documentos_personal = dp['p'] + dp['w'] + dp['i'] + dp['o']
    total_documentos_investigados = di['p'] + di['w'] + di['i'] + di['o']
    total_documentos_casos = dc['p'] + dc['w'] + dc['i'] + dc['o']
    total_documentos_direccion = dd['p'] + dd['w'] + dd['i'] + dd['o']

    # Tickets: abiertos/en_proceso primero, luego por fecha descendente
    tickets_recientes = TicketSoporte.objects.annotate(
        estado_prioridad=Case(
            When(estado='ABIERTO', then=0),
            When(estado='EN_PROCESO', then=1),
            When(estado='RESUELTO', then=2),
            When(estado='CERRADO', then=3),
            output_field=IntegerField(),
        )
    ).order_by('estado_prioridad', '-fecha_creacion')[:5]
    tickets_abiertos = TicketSoporte.objects.filter(estado__in=["ABIERTO", "EN_PROCESO"]).count()

    # Ticket counts by status — single aggregated query
    tickets_estado_qs = TicketSoporte.objects.values('estado').annotate(total=Count('id'))
    tickets_por_estado = {item['estado']: item['total'] for item in tickets_estado_qs}

    # Recent documents — single set of queries, used for both the chart list and the "by type" section
    docs_personal = DocumentoPersonal.objects.all().order_by("-fecha_subida")[:10]
    docs_investigados = DocumentoInvestigado.objects.all().order_by("-fecha_subida")[:10]
    docs_casos = DocumentoCaso.objects.all().order_by("-fecha_subida")[:10]
    docs_direccion = DocumentoDireccion.objects.all().order_by("-fecha_subida")[:10]

    all_docs_merged = sorted(
        list(docs_personal) + list(docs_investigados) + list(docs_casos) + list(docs_direccion),
        key=lambda d: d.fecha_subida,
        reverse=True
    )

    docs_pdf_list = [d for d in all_docs_merged if d.tipo == 'PDF'][:5]
    docs_word_list = [d for d in all_docs_merged if d.tipo == 'WORD'][:5]
    docs_img_list = [d for d in all_docs_merged if d.tipo == 'IMAGEN'][:5]
    docs_otro_list = [d for d in all_docs_merged if d.tipo == 'OTRO'][:5]

    # Build monthly chart data (last 6 months) — 3 queries instead of 18
    now = timezone.now()
    six_months_ago = (now - timedelta(days=180)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    month_names_es = {1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
                      7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'}

    expected_months = []
    d = six_months_ago
    for _ in range(6):
        expected_months.append((d.year, d.month))
        if d.month == 12:
            d = d.replace(year=d.year + 1, month=1)
        else:
            d = d.replace(month=d.month + 1)
    months = [month_names_es[m] for _, m in expected_months]

    def _monthly_counts(queryset):
        counts = {}
        for row in (queryset.filter(fecha_creacion__gte=six_months_ago)
                           .annotate(month=TruncMonth('fecha_creacion'))
                           .values('month')
                           .annotate(cnt=Count('id'))):
            key = (row['month'].year, row['month'].month)
            counts[key] = row['cnt']
        return [counts.get(ym, 0) for ym in expected_months]

    pm = _monthly_counts(Personal.objects)
    im = _monthly_counts(Investigado.objects)
    cm = _monthly_counts(Caso.objects)

    chart_data = json.dumps({
        "docTypes": {"pdf": docs_pdf, "word": docs_word, "img": docs_img},
        "months": months,
        "personalMonthly": pm,
        "investigadoMonthly": im,
        "casoMonthly": cm,
        "totalPersonal": total_personal,
        "totalCasos": total_casos,
        "totalInvestigados": total_investigados,
        "ticketsEstado": {
            "abierto": tickets_por_estado.get('ABIERTO', 0),
            "proceso": tickets_por_estado.get('EN_PROCESO', 0),
            "resuelto": tickets_por_estado.get('RESUELTO', 0),
            "cerrado": tickets_por_estado.get('CERRADO', 0)
        }
    })

    # Usuarios del sistema
    usuarios_sistema = User.objects.filter(is_active=True).select_related('profile').order_by('username')

    # Informes del mes actual — single query reuse for count and list
    hoy = date.today()
    informes_qs = InformeDiario.objects.filter(
        fecha__year=hoy.year, fecha__month=hoy.month
    )
    total_informes_mes = informes_qs.count()
    informes_mes = informes_qs.order_by('-fecha', '-fecha_creacion')[:10]

    # Auditoria reciente (ultimos 10 registros, si el usuario tiene permiso)
    auditoria_reciente = []
    try:
        if request.user.profile.tiene_permiso(perms.AUDITORIA_VER):
            auditoria_reciente = AuditLog.objects.all().order_by('-fecha')[:10]
    except Exception:
        pass

    context = {
        "total_personal": total_personal,
        "total_casos": total_casos,
        "total_investigados": total_investigados,
        "total_documentos_personal": total_documentos_personal,
        "total_documentos_investigados": total_documentos_investigados,
        "total_documentos_casos": total_documentos_casos,
        "total_documentos_direccion": total_documentos_direccion,
        "total_usuarios": total_usuarios,
        "usuarios_sistema": usuarios_sistema,
        "docs_pdf": docs_pdf_list,
        "docs_word": docs_word_list,
        "docs_img": docs_img_list,
        "docs_otro": docs_otro_list,
        "tareas_pendientes": tareas_pendientes,
        "tareas_count": tareas_count,
        "tareas_completadas": tareas_completadas,
        "tickets_recientes": tickets_recientes,
        "total_tickets_abiertos": tickets_abiertos,
        "tickets_abiertos_count": tickets_por_estado.get('ABIERTO', 0),
        "tickets_proceso_count": tickets_por_estado.get('EN_PROCESO', 0),
        "tickets_resueltos_count": tickets_por_estado.get('RESUELTO', 0),
        "tickets_cerrados_count": tickets_por_estado.get('CERRADO', 0),
        "chart_data_json": chart_data,
        "informes_mes": informes_mes,
        "total_informes_mes": total_informes_mes,
        "auditoria_reciente": auditoria_reciente,
        "total_bienes": total_bienes,
        "mes_actual_nombre": calendar.month_name[hoy.month].capitalize(),
    }
    return render(request, "dashboard.html", context)


# ============================================================
# USER TOGGLE
# ============================================================

@permiso_required(perms.USUARIOS_TOGGLE)
def user_toggle_active(request, pk):
    user = get_object_or_404(User, pk=pk)
    if user == request.user:
        messages.error(request, "No puedes desactivar tu propia cuenta.")
        return redirect("usuario_list")
    user.is_active = not user.is_active
    user.save()
    estado = "activada" if user.is_active else "desactivada"
    messages.success(request, f"Cuenta de {user.username} {estado} exitosamente.")
    auditar(request, "TOGGLE", "Usuario", user.pk, user.username, f"Cuenta {estado}")
    return redirect("usuario_list")


# ============================================================
# PERSONAL CRUD
# ============================================================

class PersonalListView(LoginRequiredMixin, ListView):
    model = Personal
    template_name = "direccion/personal_list.html"
    context_object_name = "personal_list"
    login_url = reverse_lazy("login")
    paginate_by = 25

    def get_queryset(self):
        queryset = Personal.objects.filter(activo=True)
        search = self.request.GET.get("search", "")
        if search:
            queryset = queryset.filter(
                Q(apellidos__icontains=search) |
                Q(nombres__icontains=search) |
                Q(cedula__icontains=search)
            )
        return queryset.order_by("apellidos", "nombres")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["search"] = self.request.GET.get("search", "")
        return context


class PersonalCreateView(PermissionRequiredMixin, CreateView):
    model = Personal
    form_class = PersonalForm
    template_name = "direccion/personal_form.html"
    login_url = reverse_lazy("login")
    permisos_requeridos = [perms.PERSONAL_CREAR]

    def get_success_url(self):
        return reverse_lazy("personal_detail", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, "Personal registrado exitosamente.")
        resp = super().form_valid(form)
        auditar(self.request, "CREAR", "Personal", self.object.pk, str(self.object), f"Cédula: {self.object.cedula}")
        return resp


class PersonalDetailView(LoginRequiredMixin, DetailView):
    model = Personal
    template_name = "direccion/personal_detail.html"
    context_object_name = "personal"
    login_url = reverse_lazy("login")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["documentos"] = DocumentoPersonal.objects.filter(personal=self.object)
        return context


class PersonalUpdateView(PermissionRequiredMixin, UpdateView):
    model = Personal
    form_class = PersonalForm
    template_name = "direccion/personal_form.html"
    login_url = reverse_lazy("login")
    permisos_requeridos = [perms.PERSONAL_EDITAR]

    def get_success_url(self):
        return reverse_lazy("personal_detail", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, "Personal actualizado exitosamente.")
        resp = super().form_valid(form)
        auditar(self.request, "ACTUALIZAR", "Personal", self.object.pk, str(self.object), f"Cédula: {self.object.cedula}")
        return resp


class PersonalDeleteView(PermissionRequiredMixin, DeleteView):
    model = Personal
    template_name = "direccion/personal_confirm_delete.html"
    success_url = reverse_lazy("personal_list")
    login_url = reverse_lazy("login")
    permisos_requeridos = [perms.PERSONAL_ELIMINAR]

    def form_valid(self, form):
        obj = self.object
        repr_ = str(obj)
        pk_val = obj.pk
        cedula = obj.cedula
        for doc in DocumentoPersonal.objects.filter(personal=obj):
            if doc.archivo:
                doc.archivo.delete(False)
        if obj.foto:
            obj.foto.delete(False)
        obj.delete()
        messages.success(self.request, f"Personal eliminado permanentemente: {repr_}")
        auditar(self.request, "ELIMINAR", "Personal", pk_val, repr_, f"Cédula: {cedula}")
        return redirect(self.success_url)


@permiso_required(perms.PERSONAL_DOCUMENTOS_AGREGAR)
def agregar_documento_personal(request, pk):
    personal = get_object_or_404(Personal, pk=pk)
    if request.method == "POST":
        form = DocumentoPersonalForm(request.POST, request.FILES)
        if form.is_valid():
            documento = form.save(commit=False)
            documento.personal = personal
            documento.save()
            messages.success(request, "Documento agregado exitosamente.")
            auditar(request, "CREAR", "DocumentoPersonal", documento.pk, str(documento), f"Personal: {personal}")
        else:
            messages.error(request, "Error al subir el documento. Verifique el formato.")
    return redirect("personal_detail", pk=pk)


@permiso_required(perms.PERSONAL_DOCUMENTOS_ELIMINAR)
def eliminar_documento_personal(request, pk, doc_pk):
    documento = get_object_or_404(DocumentoPersonal, pk=doc_pk, personal_id=pk)
    doc_repr = str(documento)
    pk_val = documento.pk
    personal_repr = str(documento.personal)
    if documento.archivo:
        documento.archivo.delete()
    documento.delete()
    messages.success(request, "Documento eliminado exitosamente.")
    auditar(request, "ELIMINAR", "DocumentoPersonal", pk_val, doc_repr, f"Personal: {personal_repr}")
    return redirect("personal_detail", pk=pk)


# ============================================================
# CASOS CRUD
# ============================================================

class CasoListView(PermissionRequiredMixin, ListView):
    model = Caso
    template_name = "direccion/caso_list.html"
    context_object_name = "casos"
    login_url = reverse_lazy("login")
    permisos_requeridos = [perms.CASOS_VER]

    def get_queryset(self):
        return Caso.objects.filter(activo=True).order_by("-fecha_creacion")


class CasoCreateView(PermissionRequiredMixin, CreateView):
    model = Caso
    form_class = CasoForm
    template_name = "direccion/caso_form.html"
    login_url = reverse_lazy("login")
    success_url = reverse_lazy("caso_list")
    permisos_requeridos = [perms.CASOS_CREAR]

    def form_valid(self, form):
        messages.success(self.request, "Caso creado exitosamente.")
        resp = super().form_valid(form)
        auditar(self.request, "CREAR", "Caso", self.object.pk, str(self.object), f"Caso: {self.object.nombre}")
        return resp


class CasoDetailView(PermissionRequiredMixin, DetailView):
    model = Caso
    template_name = "direccion/caso_detail.html"
    context_object_name = "caso"
    login_url = reverse_lazy("login")
    permisos_requeridos = [perms.CASOS_VER]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["investigados"] = Investigado.objects.filter(caso=self.object, activo=True).select_related('caso').annotate(doc_count=Count('documentos')).order_by("apellidos", "nombres")
        context["documentos"] = DocumentoCaso.objects.filter(caso=self.object)
        return context


class CasoUpdateView(PermissionRequiredMixin, UpdateView):
    model = Caso
    form_class = CasoForm
    template_name = "direccion/caso_form.html"
    login_url = reverse_lazy("login")
    success_url = reverse_lazy("caso_list")
    permisos_requeridos = [perms.CASOS_EDITAR]

    def form_valid(self, form):
        messages.success(self.request, "Caso actualizado exitosamente.")
        resp = super().form_valid(form)
        auditar(self.request, "ACTUALIZAR", "Caso", self.object.pk, str(self.object), f"Caso: {self.object.nombre}")
        return resp


class CasoDeleteView(PermissionRequiredMixin, DeleteView):
    model = Caso
    template_name = "direccion/caso_confirm_delete.html"
    success_url = reverse_lazy("caso_list")
    login_url = reverse_lazy("login")
    permisos_requeridos = [perms.CASOS_ELIMINAR]

    def form_valid(self, form):
        obj = self.object
        repr_ = str(obj)
        pk_val = obj.pk
        nombre = obj.nombre
        # Clean up associated documents from disk
        for doc in DocumentoCaso.objects.filter(caso=obj):
            if doc.archivo:
                doc.archivo.delete(False)
        Investigado.objects.filter(caso=obj).update(caso=None)
        obj.delete()
        messages.success(self.request, f"Expediente eliminado permanentemente: {repr_}")
        auditar(self.request, "ELIMINAR", "Caso", pk_val, repr_, f"Nombre: {nombre}")
        return redirect(self.success_url)


# ============================================================
# INVESTIGADOS CRUD
# ============================================================

class InvestigadoListView(PermissionRequiredMixin, ListView):
    model = Investigado
    template_name = "direccion/investigado_list.html"
    context_object_name = "investigado_list"
    login_url = reverse_lazy("login")
    paginate_by = 25
    permisos_requeridos = [perms.INVESTIGADOS_VER]

    def get_queryset(self):
        queryset = Investigado.objects.filter(activo=True).select_related("caso")
        caso_id = self.request.GET.get("caso", "")
        search = self.request.GET.get("search", "")
        if caso_id:
            queryset = queryset.filter(caso_id=caso_id)
        if search:
            queryset = queryset.filter(
                Q(apellidos__icontains=search) |
                Q(nombres__icontains=search) |
                Q(cedula__icontains=search)
            )
        return queryset.order_by("apellidos", "nombres")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["search"] = self.request.GET.get("search", "")
        context["caso_filtro"] = self.request.GET.get("caso", "")
        context["casos"] = Caso.objects.filter(activo=True).order_by("nombre")
        return context


class InvestigadoCreateView(PermissionRequiredMixin, CreateView):
    model = Investigado
    form_class = InvestigadoForm
    template_name = "direccion/investigado_form.html"
    login_url = reverse_lazy("login")
    permisos_requeridos = [perms.INVESTIGADOS_CREAR]

    def get_success_url(self):
        return reverse_lazy("investigado_detail", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, "Investigado registrado exitosamente.")
        resp = super().form_valid(form)
        auditar(self.request, "CREAR", "Investigado", self.object.pk, str(self.object), f"Cédula: {self.object.cedula or 'N/A'}")
        return resp


class InvestigadoDetailView(PermissionRequiredMixin, DetailView):
    model = Investigado
    template_name = "direccion/investigado_detail.html"
    context_object_name = "investigado"
    login_url = reverse_lazy("login")
    permisos_requeridos = [perms.INVESTIGADOS_VER]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["documentos"] = DocumentoInvestigado.objects.filter(investigado=self.object)
        return context


class InvestigadoUpdateView(PermissionRequiredMixin, UpdateView):
    model = Investigado
    form_class = InvestigadoForm
    template_name = "direccion/investigado_form.html"
    login_url = reverse_lazy("login")
    permisos_requeridos = [perms.INVESTIGADOS_EDITAR]

    def get_success_url(self):
        return reverse_lazy("investigado_detail", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, "Investigado actualizado exitosamente.")
        resp = super().form_valid(form)
        auditar(self.request, "ACTUALIZAR", "Investigado", self.object.pk, str(self.object), f"Cédula: {self.object.cedula or 'N/A'}")
        return resp


class InvestigadoDeleteView(PermissionRequiredMixin, DeleteView):
    model = Investigado
    template_name = "direccion/investigado_confirm_delete.html"
    success_url = reverse_lazy("investigado_list")
    login_url = reverse_lazy("login")
    permisos_requeridos = [perms.INVESTIGADOS_ELIMINAR]

    def form_valid(self, form):
        obj = self.object
        repr_ = str(obj)
        pk_val = obj.pk
        cedula = obj.cedula or "N/A"
        for doc in DocumentoInvestigado.objects.filter(investigado=obj):
            if doc.archivo:
                doc.archivo.delete(False)
        if obj.foto:
            obj.foto.delete(False)
        obj.delete()
        messages.success(self.request, f"Caso eliminado permanentemente: {repr_}")
        auditar(self.request, "ELIMINAR", "Investigado", pk_val, repr_, f"Cédula: {cedula}")
        return redirect(self.success_url)


@permiso_required(perms.INVESTIGADOS_DOCUMENTOS_AGREGAR)
def agregar_documento_investigado(request, pk):
    investigado = get_object_or_404(Investigado, pk=pk)
    if request.method == "POST":
        form = DocumentoInvestigadoForm(request.POST, request.FILES)
        if form.is_valid():
            documento = form.save(commit=False)
            documento.investigado = investigado
            documento.save()
            messages.success(request, "Documento agregado exitosamente.")
            auditar(request, "CREAR", "DocumentoInvestigado", documento.pk, str(documento), f"Investigado: {investigado}")
        else:
            messages.error(request, "Error al subir el documento. Verifique el formato.")
    return redirect("investigado_detail", pk=pk)


@permiso_required(perms.INVESTIGADOS_DOCUMENTOS_ELIMINAR)
def eliminar_documento_investigado(request, pk, doc_pk):
    documento = get_object_or_404(DocumentoInvestigado, pk=doc_pk, investigado_id=pk)
    doc_repr = str(documento)
    pk_val = documento.pk
    inv_repr = str(documento.investigado)
    if documento.archivo:
        documento.archivo.delete()
    documento.delete()
    messages.success(request, "Documento eliminado exitosamente.")
    auditar(request, "ELIMINAR", "DocumentoInvestigado", pk_val, doc_repr, f"Investigado: {inv_repr}")
    return redirect("investigado_detail", pk=pk)


# ============================================================
# DOCUMENTACION DE LA DIRECCION
# ============================================================

@permiso_required(perms.DOCUMENTOS_DIRECCION_VER)
def documentos_direccion_list(request):
    categoria_filtro = request.GET.get('categoria', '')
    documentos = DocumentoDireccion.objects.all().order_by('-fecha_subida')
    if categoria_filtro:
        documentos = documentos.filter(categoria=categoria_filtro)
    if request.method == "POST":
        form = DocumentoDireccionForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Documento subido exitosamente.")
            auditar(request, "CREAR", "DocumentoDireccion", form.instance.pk, str(form.instance), "Documento Direccion")
        else:
            messages.error(request, "Error al subir el documento. Verifique el formato.")
        return redirect("documentos_direccion" + (f'?categoria={categoria_filtro}' if categoria_filtro else ''))
    else:
        form = DocumentoDireccionForm()
    return render(request, "direccion/documentos_direccion.html", {
        "documentos": documentos,
        "form": form,
        "categoria_filtro": categoria_filtro,
        "categorias": DocumentoDireccion.CATEGORIA_CHOICES,
    })


@permiso_required(perms.DOCUMENTOS_DIRECCION_ELIMINAR)
def eliminar_documento_direccion(request, doc_pk):
    documento = get_object_or_404(DocumentoDireccion, pk=doc_pk)
    doc_repr = str(documento)
    pk_val = documento.pk
    if documento.archivo:
        documento.archivo.delete()
    documento.delete()
    messages.success(request, "Documento eliminado exitosamente.")
    auditar(request, "ELIMINAR", "DocumentoDireccion", pk_val, doc_repr, "Documento Direccion")
    return redirect("documentos_direccion")


# ============================================================
# CASO DOCUMENTS
# ============================================================

@permiso_required(perms.CASOS_DOCUMENTOS_AGREGAR)
def agregar_documento_caso(request, pk):
    caso = get_object_or_404(Caso, pk=pk)
    if request.method == "POST":
        form = DocumentoCasoForm(request.POST, request.FILES)
        if form.is_valid():
            documento = form.save(commit=False)
            documento.caso = caso
            documento.save()
            messages.success(request, "Documento agregado al caso exitosamente.")
            auditar(request, "CREAR", "DocumentoCaso", documento.pk, str(documento), f"Caso: {caso}")
        else:
            messages.error(request, "Error al subir el documento. Verifique el formato.")
    return redirect("caso_detail", pk=pk)


@permiso_required(perms.CASOS_DOCUMENTOS_ELIMINAR)
def eliminar_documento_caso(request, pk, doc_pk):
    documento = get_object_or_404(DocumentoCaso, pk=doc_pk, caso_id=pk)
    doc_repr = str(documento)
    pk_val = documento.pk
    caso_repr = str(documento.caso)
    if documento.archivo:
        documento.archivo.delete()
    documento.delete()
    messages.success(request, "Documento eliminado del caso exitosamente.")
    auditar(request, "ELIMINAR", "DocumentoCaso", pk_val, doc_repr, f"Caso: {caso_repr}")
    return redirect("caso_detail", pk=pk)


# ============================================================
# USER MANAGEMENT (Admin only)
# ============================================================

class UserListView(PermissionRequiredMixin, ListView):
    model = User
    template_name = "direccion/usuario_list.html"
    context_object_name = "usuarios"
    login_url = reverse_lazy("login")
    paginate_by = 25
    permisos_requeridos = [perms.USUARIOS_VER]

    def get_queryset(self):
        queryset = User.objects.all().order_by("username")
        search = self.request.GET.get("search", "")
        if search:
            queryset = queryset.filter(
                Q(username__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search)
            )
        return queryset.prefetch_related("profile")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["search"] = self.request.GET.get("search", "")
        for u in context["usuarios"]:
            try:
                u.rol = u.profile.rol
            except Exception:
                u.rol = "SIN PERFIL"
        return context


class UserCreateView(PermissionRequiredMixin, CreateView):
    model = User
    form_class = UserCreateForm
    template_name = "direccion/usuario_form.html"
    login_url = reverse_lazy("login")
    success_url = reverse_lazy("usuario_list")
    permisos_requeridos = [perms.USUARIOS_CREAR]

    def form_valid(self, form):
        messages.success(self.request, "Usuario creado exitosamente.")
        resp = super().form_valid(form)
        auditar(self.request, "CREAR", "Usuario", self.object.pk, self.object.username, f"Rol: {form.cleaned_data.get('rol', 'N/A')}")
        return resp


class UserUpdateView(PermissionRequiredMixin, UpdateView):
    model = User
    form_class = UserEditForm
    template_name = "direccion/usuario_form.html"
    login_url = reverse_lazy("login")
    success_url = reverse_lazy("usuario_list")
    permisos_requeridos = [perms.USUARIOS_EDITAR]

    def form_valid(self, form):
        messages.success(self.request, "Usuario actualizado exitosamente.")
        resp = super().form_valid(form)
        auditar(self.request, "ACTUALIZAR", "Usuario", self.object.pk, self.object.username, f"Rol: {form.cleaned_data.get('rol', 'N/A')}")
        return resp


@permiso_required(perms.USUARIOS_ELIMINAR)
def user_delete(request, pk):
    """Elimina permanentemente un usuario del sistema."""
    user = get_object_or_404(User, pk=pk)
    if user == request.user:
        messages.error(request, "No puedes eliminar tu propia cuenta.")
        return redirect("usuario_list")
    if user.is_superuser:
        messages.error(request, "No se puede eliminar el superusuario.")
        return redirect("usuario_list")
    if request.method == "POST":
        username = user.username
        pk_val = user.pk
        user.delete()
        messages.success(request, f"Usuario '{username}' eliminado permanentemente.")
        auditar(request, "ELIMINAR", "Usuario", pk_val, username, "Eliminado permanentemente")
        return redirect("usuario_list")
    return render(request, "direccion/usuario_confirm_delete.html", {"usuario": user})


# ============================================================
# EXPORT HELPERS
# ============================================================

def _build_excel_response(sheet_title, headers, rows, column_widths, filename):
    """Construye una respuesta HTTP con un archivo Excel (.xlsx) con estilo consistente."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    hf = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="003363", end_color="003363", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center")
    tb = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val).border = tb

    for col, w in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _build_pdf_response(title, headers, rows, filename):
    """Construye una respuesta HTTP con un archivo PDF con estilo corporativo."""
    LOGO_PATH = str(settings.BASE_DIR / 'static' / 'images' / 'logo.png')

    def _make_logo():
        return Image(LOGO_PATH, width=30, height=30)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter),
                            leftMargin=0.5 * inch, rightMargin=0.5 * inch,
                            topMargin=0.6 * inch, bottomMargin=0.5 * inch)
    styles = getSampleStyleSheet()

    elements = []

    # --- Header: logo + title ---
    header_cells = [
        [_make_logo(),
         Paragraph(
             f"<b>{title}</b><br/>"
             f"<font size='7.5' color='#6B7280'>Dirección General - Sistema de Gestión</font>",
             ParagraphStyle("HdrBox", parent=styles["Normal"],
                            fontSize=14, textColor=HexColor("#003363"),
                            spaceBefore=0, spaceAfter=0, leading=18)
         )]
    ]
    hdr_table = Table(header_cells, colWidths=[36, landscape(letter)[0] - inch - 36])
    hdr_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),
        ('RIGHTPADDING', (1, 0), (1, 0), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(hdr_table)

    # Separator line
    elements.append(HRFlowable(width="100%", thickness=1.5,
                                color=HexColor("#003363"),
                                spaceAfter=10, spaceBefore=4))

    # --- Data table ---
    table_data = [headers]
    for row in rows:
        table_data.append([str(c) if c is not None else "" for c in row])

    col_widths = [max(len(str(h)) * 9, 60) for h in headers]
    total_w = sum(col_widths)
    page_w = landscape(letter)[0] - inch
    if total_w > page_w:
        col_widths = [w * page_w / total_w for w in col_widths]

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#003363")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.Color(0, 0, 0, 0.2)),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.Color(1, 1, 1, 1), colors.Color(0.95, 0.95, 0.97, 1)]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 12))

    # Footer
    foot_style = ParagraphStyle(
        "Foot", parent=styles["Normal"],
        fontSize=7, textColor=HexColor("#9CA3AF"), alignment=1
    )
    elements.append(Paragraph(
        f"Generado: {datetime.now():%d/%m/%Y %H:%M} | Total: {len(rows)} registros",
        foot_style
    ))

    doc.build(elements)
    pdf = buf.getvalue()
    buf.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _human_size(size_bytes):
    """Convierte bytes a formato legible (KB, MB, GB)."""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024:
            return f'{size_bytes:.1f} {unit}'
        size_bytes /= 1024
    return f'{size_bytes:.1f} TB'


# ============================================================
# EXPORT TO EXCEL
# ============================================================

@permiso_required(perms.EXPORTAR_PERSONAL_EXCEL)
def exportar_personal_excel(request):
    headers = ["Apellidos", "Nombres", "Cedula", "Telefonos", "Fecha Nacimiento", "Fecha Registro"]
    rows = []
    for p in Personal.objects.filter(activo=True).order_by("apellidos"):
        rows.append([
            p.apellidos,
            p.nombres,
            p.cedula,
            p.telefonos or "",
            p.fecha_nacimiento.strftime("%d/%m/%Y") if p.fecha_nacimiento else "",
            p.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
        ])
    auditar(request, "EXPORTAR", "Personal", None, "Exportación Excel",
            f"{Personal.objects.filter(activo=True).count()} registros")
    return _build_excel_response(
        "Personal", headers, rows,
        [30, 30, 18, 30, 18, 20], "personal_direccion.xlsx"
    )


@permiso_required(perms.EXPORTAR_INVESTIGADOS_EXCEL)
def exportar_investigados_excel(request):
    headers = ["Apellidos", "Nombres", "Cedula", "RIF", "Partida Nacimiento", "Entrada Investigacion", "Fecha Registro"]
    rows = []
    for i in Investigado.objects.filter(activo=True).order_by("apellidos"):
        rows.append([
            i.apellidos,
            i.nombres,
            i.cedula or "",
            i.rif or "",
            i.partida_nacimiento or "",
            i.entrada_investigacion or "",
            i.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
        ])
    auditar(request, "EXPORTAR", "Investigado", None, "Exportación Excel",
            f"{Investigado.objects.filter(activo=True).count()} registros")
    return _build_excel_response(
        "Investigados", headers, rows,
        [30, 30, 18, 18, 25, 40, 20], "investigados.xlsx"
    )


# ============================================================
# EXPORT TO PDF
# ============================================================

@permiso_required(perms.EXPORTAR_PERSONAL_PDF)
def exportar_personal_pdf(request):
    qs = Personal.objects.filter(activo=True).order_by("apellidos", "nombres")
    headers = ["Apellidos", "Nombres", "Cédula", "Teléfonos", "Fecha Nac.", "Fecha Registro"]
    rows = []
    for p in qs:
        rows.append([
            p.apellidos,
            p.nombres,
            p.cedula,
            p.telefonos or "",
            p.fecha_nacimiento.strftime("%d/%m/%Y") if p.fecha_nacimiento else "",
            p.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
        ])
    auditar(request, "EXPORTAR", "Personal", None, "Exportación PDF", f"{qs.count()} registros")
    return _build_pdf_response(
        "Personal de la Dirección General - Reporte",
        headers, rows, "personal_direccion.pdf"
    )


@permiso_required(perms.EXPORTAR_INVESTIGADOS_PDF)
def exportar_investigados_pdf(request):
    qs = Investigado.objects.filter(activo=True).order_by("apellidos", "nombres")
    headers = ["Apellidos", "Nombres", "Cédula", "RIF", "Partida Nacimiento", "Entrada Investigación", "Fecha Registro"]
    rows = []
    for i in qs:
        rows.append([
            i.apellidos,
            i.nombres,
            i.cedula or "",
            i.rif or "",
            i.partida_nacimiento or "",
            i.entrada_investigacion or "",
            i.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
        ])
    auditar(request, "EXPORTAR", "Investigado", None, "Exportación PDF", f"{qs.count()} registros")
    return _build_pdf_response(
        "Personas a Investigar - Reporte",
        headers, rows, "investigados.pdf"
    )


# ============================================================
# NOTIFICACIONES
# ============================================================

@permiso_required(perms.NOTIFICACIONES_VER)
def notificaciones_list(request):
    """Lista todas las notificaciones y marca las no leídas como leídas."""
    notis = Notificacion.objects.filter(usuario=request.user)
    notis.filter(leida=False).update(leida=True)
    # Invalidar cache del badge de notificaciones
    from django.core.cache import cache
    cache.delete(f'notis_unread_{request.user.pk}')
    return render(request, "direccion/notificaciones_list.html", {
        "notificaciones": notis,
    })


# ============================================================
# TAREAS PENDIENTES (any role)
# ============================================================

@permiso_required(perms.TAREAS_VER)
def tareas_list(request):
    """Lista todas las tareas pendientes. Cualquier usuario puede añadir."""
    if request.method == "POST":
        form = TareaForm(request.POST)
        if form.is_valid():
            tarea = form.save(commit=False)
            tarea.creado_por = request.user
            tarea.save()
            messages.success(request, "Tarea agregada.")
        return redirect("tareas_list")
    tareas = Tarea.objects.all()
    form = TareaForm()
    return render(request, "direccion/tareas_list.html", {
        "tareas": tareas,
        "form": form,
    })


@permiso_required(perms.TAREAS_COMPLETAR)
def tarea_completar(request, pk):
    """Marca/desmarca una tarea como completada."""
    tarea = get_object_or_404(Tarea, pk=pk)
    tarea.completada = not tarea.completada
    tarea.save()
    estado = "completada" if tarea.completada else "pendiente"
    messages.success(request, f"Tarea marcada como {estado}.")
    return redirect("tareas_list")


@permiso_required(perms.TAREAS_ELIMINAR)
def tarea_eliminar(request, pk):
    """Elimina una tarea."""
    tarea = get_object_or_404(Tarea, pk=pk)
    tarea.delete()
    messages.success(request, "Tarea eliminada.")
    return redirect("tareas_list")


# ============================================================
# TICKETS DE SOPORTE
# ============================================================

@permiso_required(perms.TICKETS_VER)
def ticket_detail(request, pk):
    """Detalle de un ticket con su historial de cambios."""
    ticket = get_object_or_404(TicketSoporte.objects.select_related('creado_por', 'asignado_a'), pk=pk)
    return render(request, "direccion/ticket_detail.html", {
        "ticket": ticket,
        "historial": ticket.historial.all(),
    })


@permiso_required(perms.TICKETS_VER)
def ticket_list(request):
    """Lista de tickets con filtros por estado y usuario."""
    filtro_estado = request.GET.get("estado", "")
    filtro_usuario = request.GET.get("usuario", "")

    # Admin/Supervisor ven todos los tickets; el resto solo los propios
    if request.user.profile.tiene_permiso(perms.TICKETS_RESOLVER):
        tickets = TicketSoporte.objects.select_related('creado_por', 'asignado_a').all()
        if filtro_usuario:
            tickets = tickets.filter(creado_por_id=filtro_usuario)
    else:
        tickets = TicketSoporte.objects.select_related('creado_por', 'asignado_a').filter(creado_por=request.user)

    if filtro_estado:
        tickets = tickets.filter(estado=filtro_estado)

    # Orden: abiertos/en_proceso primero, luego por fecha descendente
    tickets = tickets.annotate(
        estado_prioridad=Case(
            When(estado='ABIERTO', then=0),
            When(estado='EN_PROCESO', then=1),
            When(estado='RESUELTO', then=2),
            When(estado='CERRADO', then=3),
            output_field=IntegerField(),
        )
    ).order_by('estado_prioridad', '-fecha_creacion')

    usuarios = User.objects.filter(is_active=True).order_by("username") if request.user.profile.tiene_permiso(perms.TICKETS_RESOLVER) else []

    return render(request, "direccion/ticket_list.html", {
        "tickets": tickets,
        "filtro_estado": filtro_estado,
        "filtro_usuario": filtro_usuario,
        "usuarios": usuarios,
        "estados": TicketSoporte.ESTADO_CHOICES,
    })


@permiso_required(perms.TICKETS_CREAR)
def ticket_create(request):
    """Crea un nuevo ticket de soporte."""
    if request.method == "POST":
        form = TicketSoporteForm(request.POST)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.creado_por = request.user
            ticket.save()
            # Registrar creación en el historial
            ticket.registrar_cambio(request.user, 'estado', '---', 'ABIERTO')
            ticket.registrar_cambio(request.user, 'prioridad', '---', ticket.prioridad)
            messages.success(request, f"Ticket #{ticket.pk} creado exitosamente.")
            auditar(request, "CREAR", "TicketSoporte", ticket.pk, str(ticket), f"Asunto: {ticket.asunto}")
            notificar_administradores(
                f"Nuevo ticket #{ticket.pk}: {ticket.asunto} (por {request.user.get_full_name() or request.user.username})",
                link="/tickets/"
            )
            return redirect("ticket_detail", pk=ticket.pk)
    else:
        form = TicketSoporteForm()
    return render(request, "direccion/ticket_form.html", {
        "form": form,
        "accion": "Crear"
    })


@permiso_required(perms.TICKETS_RESOLVER)
def ticket_resolver(request, pk):
    """Marca un ticket como resuelto (un solo clic desde la lista)."""
    ticket = get_object_or_404(TicketSoporte, pk=pk)
    estado_anterior = ticket.estado
    if ticket.estado == 'RESUELTO':
        ticket.estado = 'ABIERTO'
        mensaje = "reabierto"
    else:
        ticket.estado = 'RESUELTO'
        mensaje = "resuelto"
    ticket.save()
    ticket.registrar_cambio(request.user, 'estado', estado_anterior, ticket.estado)
    messages.success(request, f"Ticket #{ticket.pk} marcado como {mensaje}.")
    auditar(request, "ACTUALIZAR", "TicketSoporte", ticket.pk, str(ticket),
            f"Estado: {ticket.get_estado_display()}")
    return redirect("ticket_list")


@permiso_required(perms.TICKETS_ASIGNAR)
def ticket_asignar(request, pk):
    """Asigna un ticket a un usuario y/o cambia su estado (solo admin)."""
    ticket = get_object_or_404(TicketSoporte, pk=pk)
    estado_anterior = ticket.estado
    prioridad_anterior = ticket.prioridad
    asignado_anterior = ticket.asignado_a

    if request.method == "POST":
        form = TicketAsignarForm(request.POST, instance=ticket)
        if form.is_valid():
            form.save()
            ticket.registrar_cambio(request.user, 'estado', estado_anterior, ticket.estado)
            ticket.registrar_cambio(request.user, 'prioridad', prioridad_anterior, ticket.prioridad)
            anterior_nombre = asignado_anterior.get_full_name() or asignado_anterior.username if asignado_anterior else '---'
            nuevo_nombre = ticket.asignado_a.get_full_name() or ticket.asignado_a.username if ticket.asignado_a else '---'
            ticket.registrar_cambio(request.user, 'asignado_a', anterior_nombre, nuevo_nombre)
            messages.success(request, f"Ticket #{ticket.pk} actualizado.")
            auditar(request, "ACTUALIZAR", "TicketSoporte", ticket.pk, str(ticket),
                    f"Estado: {ticket.get_estado_display()}, Asignado: {ticket.asignado_a or 'Nadie'}")
            return redirect("ticket_detail", pk=ticket.pk)
    else:
        form = TicketAsignarForm(instance=ticket)
    return render(request, "direccion/ticket_form.html", {
        "form": form,
        "ticket": ticket,
        "accion": "Asignar"
    })


# ============================================================
# INFORMES DIARIOS
# ============================================================


def _generar_pdf_informe(informe):
    """Genera un PDF en memoria para un InformeDiario y devuelve los bytes."""
    pdf_buf = BytesIO()
    styles = getSampleStyleSheet()

    doc = SimpleDocTemplate(
        pdf_buf, pagesize=letter,
        leftMargin=0.75 * inch, rightMargin=0.75 * inch,
        topMargin=0.6 * inch, bottomMargin=0.6 * inch
    )

    elements = []

    # Título
    elements.append(Paragraph(
        f"<b>{informe.titulo}</b>",
        ParagraphStyle("TitlePDF", parent=styles["Normal"],
                       fontSize=16, textColor=HexColor("#003363"),
                       spaceAfter=6, leading=20)
    ))

    # Fecha y autor
    autor = informe.creado_por.get_full_name() or informe.creado_por.username if informe.creado_por else 'Sistema'
    elements.append(Paragraph(
        f"<font size='9' color='#6B7280'>Fecha: {informe.fecha.strftime('%d/%m/%Y')} &nbsp;|&nbsp; Autor: {autor}</font>",
        ParagraphStyle("MetaPDF", parent=styles["Normal"],
                       fontSize=9, textColor=HexColor("#6B7280"),
                       spaceAfter=8, leading=12)
    ))

    # Línea separadora
    elements.append(HRFlowable(width="100%", thickness=1, color=HexColor("#E5E7EB"), spaceAfter=14, spaceBefore=2))

    # Contenido
    contenido_html = informe.contenido.replace('\n', '<br/>')
    elements.append(Paragraph(
        contenido_html,
        ParagraphStyle("BodyPDF", parent=styles["Normal"],
                       fontSize=10, leading=15, spaceAfter=10)
    ))

    # Archivo adjunto
    if informe.archivo:
        elements.append(Spacer(1, 12))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=HexColor("#D1D5DB"), spaceAfter=8, spaceBefore=2))
        elements.append(Paragraph(
            f"<font size='9' color='#003363'>Archivo adjunto: {informe.archivo.name.split('/')[-1]}</font>",
            ParagraphStyle("AttachPDF", parent=styles["Normal"],
                           fontSize=9, textColor=HexColor("#003363"),
                           spaceAfter=6, leading=12)
        ))

    # Footer
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        f"<font size='7.5' color='#9CA3AF'>Generado: {datetime.now():%d/%m/%Y %H:%M}</font>",
        ParagraphStyle("FooterPDF", parent=styles["Normal"],
                       fontSize=7.5, textColor=HexColor("#9CA3AF"),
                       alignment=1)
    ))

    doc.build(elements)
    pdf_data = pdf_buf.getvalue()
    pdf_buf.close()
    return pdf_data


@permiso_required(perms.INFORMES_PREVIEW)
def previsualizar_informe_pdf(request, pk):
    """Muestra un PDF inline de un informe diario individual para previsualización."""
    informe = get_object_or_404(InformeDiario, pk=pk)
    pdf_data = _generar_pdf_informe(informe)

    safe_titulo = informe.titulo.replace(' ', '_')[:60]
    filename = f"{informe.fecha}_{safe_titulo}.pdf"

    response = HttpResponse(pdf_data, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


@permiso_required(perms.INFORMES_DESCARGAR)
def exportar_informes_descargar(request):
    """Genera un ZIP con uno o varios informes diarios."""
    tipo = request.GET.get('tipo', 'mes')
    pk = request.GET.get('pk', '')
    semana = request.GET.get('semana', '')
    mes = request.GET.get('mes', '')
    anio = request.GET.get('anio', '')
    hoy = date.today()

    if not anio:
        anio = str(hoy.year)

    informes = InformeDiario.objects.none()
    nombre_zip = "informes.zip"

    if tipo == 'semana' and semana:
        try:
            semana_int = int(semana)
            from datetime import timedelta
            primer_dia = date(int(anio), 1, 1)
            dias_restar = primer_dia.weekday()
            if dias_restar <= 3:
                lunes_sem1 = primer_dia - timedelta(days=dias_restar)
            else:
                lunes_sem1 = primer_dia + timedelta(days=(7 - dias_restar))
            lunes = lunes_sem1 + timedelta(weeks=semana_int - 1)
            domingo = lunes + timedelta(days=6)
            informes = InformeDiario.objects.filter(fecha__gte=lunes, fecha__lte=domingo).order_by('fecha')
        except (ValueError, TypeError):
            pass
        nombre_zip = f"informes_semana_{semana}_{anio}.zip"
    else:  # mes
        if not mes:
            mes = str(hoy.month)
        informes = InformeDiario.objects.filter(
            fecha__year=int(anio), fecha__month=int(mes)
        ).order_by('fecha')
        nombre_mes = calendar.month_name[int(mes)].capitalize()
        nombre_zip = f"informes_{nombre_mes.lower()}_{anio}.zip"

    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for informe in informes:
            safe_titulo = re.sub(r'[^\w\s-]', '', informe.titulo)
            safe_titulo = re.sub(r'[-\s]+', '_', safe_titulo).strip('_')[:60]
            pdf_filename = f"{informe.fecha}_{safe_titulo}.pdf"

            pdf_data = _generar_pdf_informe(informe)
            zf.writestr(pdf_filename, pdf_data)

    zip_data = buf.getvalue()
    buf.close()

    auditar(request, "EXPORTAR", "InformeDiario", None,
            f"ZIP {nombre_zip}", f"{informes.count()} informe(s)")

    response = HttpResponse(zip_data, content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="{nombre_zip}"'
    return response


@permiso_required(perms.INFORMES_ELIMINAR)
def eliminar_informe_diario(request, pk):
    """Elimina un informe diario."""
    informe = get_object_or_404(InformeDiario, pk=pk)
    if request.method == "POST":
        repr_ = str(informe)
        pk_val = informe.pk
        if informe.archivo:
            informe.archivo.delete()
        informe.delete()
        messages.success(request, f"Informe '{repr_}' eliminado exitosamente.")
        auditar(request, "ELIMINAR", "InformeDiario", pk_val, repr_,
                f"Título: {repr_}")
        return redirect("informes_diarios")
    return redirect("informes_diarios")


@permiso_required(perms.INFORMES_VER)
def informes_diarios_list(request):
    """Lista de informes diarios con filtros por semana, mes y año."""
    periodo = request.GET.get('periodo', 'mes')
    anio = request.GET.get('anio', '')
    mes = request.GET.get('mes', '')
    semana = request.GET.get('semana', '')

    informes = InformeDiario.objects.all()
    hoy = date.today()

    # Valores por defecto
    if not anio:
        anio = str(hoy.year)
    if not mes:
        mes = str(hoy.month)

    # Aplicar filtros
    informes = informes.filter(fecha__year=int(anio))

    if periodo == 'semana' and semana:
        try:
            semana_int = int(semana)
            # Calcular rango de fechas para la semana ISO
            from datetime import timedelta
            # Primero obtenemos el primer día del año
            primer_dia = date(int(anio), 1, 1)
            # Ajustar al lunes de la semana 1 ISO
            dias_restar = primer_dia.weekday()  # 0 = lunes
            if dias_restar <= 3:
                lunes_sem1 = primer_dia - timedelta(days=dias_restar)
            else:
                lunes_sem1 = primer_dia + timedelta(days=(7 - dias_restar))
            lunes = lunes_sem1 + timedelta(weeks=semana_int - 1)
            domingo = lunes + timedelta(days=6)
            informes = informes.filter(fecha__gte=lunes, fecha__lte=domingo)
        except (ValueError, TypeError):
            pass

    if periodo == 'mes' and mes:
        informes = informes.filter(fecha__month=int(mes))

    # Generar opciones para los filtros
    anios_disponibles = InformeDiario.objects.dates('fecha', 'year', order='DESC')
    if not anios_disponibles:
        anios_disponibles = [hoy]

    # Meses en español
    meses_opciones = []
    for i in range(1, 13):
        meses_opciones.append((i, calendar.month_name[i].capitalize()))

    # Gestión de creación/edición
    if request.method == "POST":
        form = InformeDiarioForm(request.POST, request.FILES)
        if form.is_valid():
            informe = form.save(commit=False)
            informe.creado_por = request.user
            informe.save()
            messages.success(request, f"Informe '{informe.titulo}' creado exitosamente.")
            auditar(request, "CREAR", "InformeDiario", informe.pk, str(informe), f"Título: {informe.titulo}")
        else:
            messages.error(request, "Error al crear el informe. Verifica los datos.")
        return redirect(request.path + (f'?periodo={periodo}&anio={anio}&mes={mes}&semana={semana}' if periodo else ''))
    else:
        form = InformeDiarioForm()

    return render(request, "direccion/informes_diarios.html", {
        "informes": informes,
        "form": form,
        "periodo": periodo,
        "anio_actual": anio,
        "mes_actual": mes,
        "semana_actual": semana,
        "anios_disponibles": anios_disponibles,
        "meses_opciones": meses_opciones,
        "hoy": hoy,

    })


# ============================================================
# BÚSQUEDA GLOBAL
# ============================================================

@permiso_required(perms.DASHBOARD_VER)
def busqueda_global(request):
    """Endpoint JSON para búsqueda global en todos los módulos del sistema.
    Retorna resultados agrupados por modelo, ordenados por relevancia.
    Filtra según los permisos del usuario autenticado.
    """
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'results': {}})

    try:
        profile = request.user.profile
    except Exception:
        return JsonResponse({'results': {}})

    results = {}

    # Personal
    if profile.tiene_permiso(perms.PERSONAL_VER):
        personal_qs = Personal.objects.filter(activo=True).filter(
            Q(apellidos__icontains=q) |
            Q(nombres__icontains=q) |
            Q(cedula__icontains=q)
        )[:8]
        if personal_qs:
            results['personal'] = {
                'label': 'Personal',
                'icon': 'users',
                'items': [{
                    'id': p.pk,
                    'title': f"{p.apellidos}, {p.nombres}",
                    'subtitle': p.cedula,
                    'url': reverse_lazy('personal_detail', args=[p.pk]),
                } for p in personal_qs]
            }

    # Casos
    if profile.tiene_permiso(perms.CASOS_VER):
        casos_qs = Caso.objects.filter(activo=True).filter(
            Q(nombre__icontains=q) |
            Q(descripcion__icontains=q)
        )[:5]
        if casos_qs:
            results['casos'] = {
                'label': 'Expedientes',
                'icon': 'folder',
                'items': [{
                    'id': c.pk,
                    'title': c.nombre,
                    'subtitle': c.descripcion or '',
                    'url': reverse_lazy('caso_detail', args=[c.pk]),
                } for c in casos_qs]
            }

    # Investigados
    if profile.tiene_permiso(perms.INVESTIGADOS_VER):
        inv_qs = Investigado.objects.filter(activo=True).filter(
            Q(apellidos__icontains=q) |
            Q(nombres__icontains=q) |
            Q(cedula__icontains=q)
        )[:8]
        if inv_qs:
            results['investigados'] = {
                'label': 'Investigados',
                'icon': 'search',
                'items': [{
                    'id': i.pk,
                    'title': f"{i.apellidos}, {i.nombres}",
                    'subtitle': i.cedula or 'Sin cédula',
                    'url': reverse_lazy('investigado_detail', args=[i.pk]),
                } for i in inv_qs]
            }

    # Documentos de la Dirección
    if profile.tiene_permiso(perms.DOCUMENTOS_DIRECCION_VER):
        docs_qs = DocumentoDireccion.objects.filter(
            Q(descripcion__icontains=q)
        )[:5]
        if docs_qs:
            results['documentos'] = {
                'label': 'Documentación',
                'icon': 'file',
                'items': [{
                    'id': d.pk,
                    'title': d.descripcion or d.archivo.name.split('/')[-1],
                    'subtitle': f"{d.get_categoria_display()} · {d.get_tipo_display()}",
                    'url': d.archivo.url if d.archivo else '#',
                } for d in docs_qs]
            }

    # Tickets de Soporte
    if profile.tiene_permiso(perms.TICKETS_VER):
        tickets_qs = TicketSoporte.objects.filter(
            Q(asunto__icontains=q) |
            Q(descripcion__icontains=q)
        ).select_related('creado_por')[:5]
        if tickets_qs:
            results['tickets'] = {
                'label': 'Tickets de Soporte',
                'icon': 'ticket',
                'items': [{
                    'id': t.pk,
                    'title': f"#{t.pk} - {t.asunto}",
                    'subtitle': f"{t.get_estado_display()} · {t.creado_por.get_full_name() or t.creado_por.username if t.creado_por else 'Sistema'}",
                    'url': reverse_lazy('ticket_detail', args=[t.pk]),
                } for t in tickets_qs]
            }

    # Tareas
    if profile.tiene_permiso(perms.TAREAS_VER):
        tareas_qs = Tarea.objects.filter(
            Q(descripcion__icontains=q)
        )[:5]
        if tareas_qs:
            results['tareas'] = {
                'label': 'Tareas',
                'icon': 'checklist',
                'items': [{
                    'id': t.pk,
                    'title': t.descripcion[:80],
                    'subtitle': f"{'✓ Completada' if t.completada else '○ Pendiente'} · {t.get_prioridad_display()}",
                    'url': reverse_lazy('tareas_list'),
                } for t in tareas_qs]
            }

    # Bienes
    if profile.tiene_permiso(perms.BIENES_VER):
        bienes_qs = Bien.objects.filter(activo=True).filter(
            Q(nombre__icontains=q) |
            Q(codigo_inventario__icontains=q) |
            Q(serial__icontains=q) |
            Q(marca__icontains=q) |
            Q(ubicacion__icontains=q)
        )[:5]
        if bienes_qs:
            results['bienes'] = {
                'label': 'Bienes',
                'icon': 'folder',
                'items': [{
                    'id': b.pk,
                    'title': f"{b.nombre}",
                    'subtitle': f"{b.get_categoria_display()} · {b.get_estado_display()}",
                    'url': reverse_lazy('bien_detail', args=[b.pk]),
                } for b in bienes_qs]
            }

    # Informes Diarios
    if profile.tiene_permiso(perms.INFORMES_VER):
        informes_qs = InformeDiario.objects.filter(
            Q(titulo__icontains=q) |
            Q(contenido__icontains=q)
        )[:5]
        if informes_qs:
            results['informes'] = {
                'label': 'Informes Diarios',
                'icon': 'chart',
                'items': [{
                    'id': i.pk,
                    'title': i.titulo,
                    'subtitle': i.fecha.strftime('%d/%m/%Y'),
                    'url': reverse_lazy('informe_diario_preview', args=[i.pk]),
                } for i in informes_qs]
            }

    return JsonResponse({'results': results, 'total': sum(len(v['items']) for v in results.values())})


# ============================================================
# BIENES CRUD
# ============================================================

class BienListView(PermissionRequiredMixin, ListView):
    model = Bien
    template_name = "direccion/bien_list.html"
    context_object_name = "bienes"
    login_url = reverse_lazy("login")
    paginate_by = 25
    permisos_requeridos = [perms.BIENES_VER]

    def get_queryset(self):
        queryset = Bien.objects.filter(activo=True)
        search = self.request.GET.get("search", "")
        categoria = self.request.GET.get("categoria", "")
        if search:
            queryset = queryset.filter(
                Q(nombre__icontains=search) |
                Q(codigo_inventario__icontains=search) |
                Q(serial__icontains=search) |
                Q(marca__icontains=search) |
                Q(ubicacion__icontains=search)
            )
        if categoria:
            queryset = queryset.filter(categoria=categoria)
        return queryset.order_by("-fecha_creacion")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["search"] = self.request.GET.get("search", "")
        context["categoria_filtro"] = self.request.GET.get("categoria", "")
        context["categorias"] = Bien.CATEGORIA_CHOICES
        return context


class BienCreateView(PermissionRequiredMixin, CreateView):
    model = Bien
    form_class = BienForm
    template_name = "direccion/bien_form.html"
    login_url = reverse_lazy("login")
    permisos_requeridos = [perms.BIENES_CREAR]

    def get_success_url(self):
        return reverse_lazy("bien_detail", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, "Bien registrado exitosamente.")
        resp = super().form_valid(form)
        auditar(self.request, "CREAR", "Bien", self.object.pk, str(self.object), f"Nombre: {self.object.nombre}")
        return resp


class BienDetailView(PermissionRequiredMixin, DetailView):
    model = Bien
    template_name = "direccion/bien_detail.html"
    context_object_name = "bien"
    login_url = reverse_lazy("login")
    permisos_requeridos = [perms.BIENES_VER]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["documentos"] = DocumentoBien.objects.filter(bien=self.object)
        return context


class BienUpdateView(PermissionRequiredMixin, UpdateView):
    model = Bien
    form_class = BienForm
    template_name = "direccion/bien_form.html"
    login_url = reverse_lazy("login")
    permisos_requeridos = [perms.BIENES_EDITAR]

    def get_success_url(self):
        return reverse_lazy("bien_detail", kwargs={"pk": self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, "Bien actualizado exitosamente.")
        resp = super().form_valid(form)
        auditar(self.request, "ACTUALIZAR", "Bien", self.object.pk, str(self.object), f"Nombre: {self.object.nombre}")
        return resp


class BienDeleteView(PermissionRequiredMixin, DeleteView):
    model = Bien
    template_name = "direccion/bien_confirm_delete.html"
    success_url = reverse_lazy("bien_list")
    login_url = reverse_lazy("login")
    permisos_requeridos = [perms.BIENES_ELIMINAR]

    def form_valid(self, form):
        obj = self.object
        repr_ = str(obj)
        pk_val = obj.pk
        nombre = obj.nombre
        for doc in DocumentoBien.objects.filter(bien=obj):
            if doc.archivo:
                doc.archivo.delete(False)
        if obj.foto:
            obj.foto.delete(False)
        obj.delete()
        messages.success(self.request, f"Bien eliminado permanentemente: {repr_}")
        auditar(self.request, "ELIMINAR", "Bien", pk_val, repr_, f"Nombre: {nombre}")
        return redirect(self.success_url)


@permiso_required(perms.BIENES_DOCUMENTOS_AGREGAR)
def agregar_documento_bien(request, pk):
    bien = get_object_or_404(Bien, pk=pk)
    if request.method == "POST":
        form = DocumentoBienForm(request.POST, request.FILES)
        if form.is_valid():
            documento = form.save(commit=False)
            documento.bien = bien
            documento.save()
            messages.success(request, "Documento agregado exitosamente.")
            auditar(request, "CREAR", "DocumentoBien", documento.pk, str(documento), f"Bien: {bien}")
        else:
            messages.error(request, "Error al subir el documento. Verifique el formato.")
    return redirect("bien_detail", pk=pk)


@permiso_required(perms.BIENES_DOCUMENTOS_ELIMINAR)
def eliminar_documento_bien(request, pk, doc_pk):
    documento = get_object_or_404(DocumentoBien, pk=doc_pk, bien_id=pk)
    doc_repr = str(documento)
    pk_val = documento.pk
    bien_repr = str(documento.bien)
    if documento.archivo:
        documento.archivo.delete()
    documento.delete()
    messages.success(request, "Documento eliminado exitosamente.")
    auditar(request, "ELIMINAR", "DocumentoBien", pk_val, doc_repr, f"Bien: {bien_repr}")
    return redirect("bien_detail", pk=pk)


# ============================================================
# REPORTES - PANEL CENTRALIZADO
# ============================================================

@permiso_required(perms.REPORTES_VER)
def panel_reportes(request):
    """Panel central de reportes con filtros y descarga."""
    fec_desde = request.GET.get("fecha_desde", "")
    fec_hasta = request.GET.get("fecha_hasta", "")

    q_personal = Personal.objects.filter(activo=True)
    q_investigados = Investigado.objects.filter(activo=True)
    q_docs_p = DocumentoPersonal.objects.all()
    q_docs_i = DocumentoInvestigado.objects.all()
    q_docs_c = DocumentoCaso.objects.all()

    if fec_desde:
        try:
            desde_dt = datetime.strptime(fec_desde, "%Y-%m-%d")
            q_personal = q_personal.filter(fecha_creacion__gte=desde_dt)
            q_investigados = q_investigados.filter(fecha_creacion__gte=desde_dt)
        except (ValueError, TypeError):
            pass
    if fec_hasta:
        try:
            hasta_dt = datetime.strptime(fec_hasta, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            q_personal = q_personal.filter(fecha_creacion__lte=hasta_dt)
            q_investigados = q_investigados.filter(fecha_creacion__lte=hasta_dt)
        except (ValueError, TypeError):
            pass

    context = {
        "total_personal": q_personal.count(),
        "total_investigados": q_investigados.count(),
        "total_documentos": q_docs_p.count() + q_docs_i.count() + q_docs_c.count(),
        "fecha_desde": fec_desde,
        "fecha_hasta": fec_hasta,
    }
    return render(request, "reportes.html", context)


# ============================================================
# BACKUP & RESTORE (Admin only)
# ============================================================

@permiso_required(perms.BACKUP_DESCARGAR)
def backup_view(request):
    """Panel de copias de seguridad: descargar backup o restaurar.
    Compatible con SQLite y PostgreSQL (usa dumpdata/loaddata).
    """
    db_engine = settings.DATABASES['default']['ENGINE']
    db_name = settings.DATABASES['default']['NAME']

    # Obtener info del tamaño de BD según el motor
    db_size = 0
    if 'sqlite' in db_engine:
        db_path = Path(settings.DATABASES['default']['NAME'])
        db_size = db_path.stat().st_size if db_path.exists() else 0

    media_root = settings.MEDIA_ROOT
    media_count = sum(1 for _ in media_root.rglob('*') if _.is_file()) if media_root.exists() else 0
    media_size = sum(f.stat().st_size for f in media_root.rglob('*') if f.is_file()) if media_root.exists() else 0

    context = {
        'db_size': db_size,
        'db_size_human': _human_size(db_size) if 'sqlite' in db_engine else 'Gestionado por PostgreSQL',
        'media_count': media_count,
        'media_size_human': _human_size(media_size),
        'total_size_human': _human_size(db_size + media_size) if 'sqlite' in db_engine else _human_size(media_size),
        'fecha_actual': datetime.now().strftime('%d/%m/%Y %H:%M'),
        'db_engine': 'PostgreSQL' if 'postgresql' in db_engine else 'SQLite',
    }
    return render(request, 'direccion/backup.html', context)


@permiso_required(perms.BACKUP_DESCARGAR)
def descargar_backup(request):
    """Genera y descarga un archivo ZIP con dump de la BD y los archivos media.
    Compatible con SQLite y PostgreSQL (usa dumpdata de Django).
    """
    from django.core.management import call_command

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)

        # 1. Backup de la base de datos usando dumpdata (funciona con cualquier motor)
        backup_json = tmp / 'backup.json'
        with open(str(backup_json), 'w', encoding='utf-8') as f:
            call_command('dumpdata', '--natural-foreign', '--natural-primary',
                         '--exclude', 'contenttypes', '--exclude', 'auth.permission',
                         stdout=f)

        # 2. Copiar archivos media
        media_copy = tmp / 'media'
        if settings.MEDIA_ROOT.exists():
            shutil.copytree(str(settings.MEDIA_ROOT), str(media_copy), dirs_exist_ok=True)

        # 3. Crear ZIP
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        zip_name = f'respaldo_{timestamp}.zip'
        zip_path = tmp / zip_name

        with zipfile.ZipFile(str(zip_path), 'w', zipfile.ZIP_DEFLATED) as zf:
            if backup_json.exists():
                zf.write(str(backup_json), 'backup.json')
            if media_copy.exists():
                for f in sorted(media_copy.rglob('*')):
                    if f.is_file():
                        rel = f.relative_to(media_copy)
                        zf.write(str(f), f'media/{rel}')

        # 4. Información del backup
        with zipfile.ZipFile(str(zip_path), 'r') as zf:
            info = zf.infolist()

        auditar(request, "BACKUP", "Sistema", None, "Descarga de respaldo",
                f"{zip_name} - {len(info)} archivos")

        # 5. Servir el archivo
        response = HttpResponse(content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{zip_name}"'
        with open(str(zip_path), 'rb') as f:
            response.write(f.read())
        return response


@permiso_required(perms.BACKUP_RESTAURAR)
def restaurar_backup(request):
    """Recibe un archivo ZIP de respaldo y restaura la BD y media.
    Compatible con SQLite y PostgreSQL (usa dumpdata/loaddata de Django).
    Soporta tanto el formato nuevo (backup.json) como el antiguo (db.sqlite3).
    """
    from django.core.management import call_command

    if request.method != 'POST':
        return redirect('backup')

    archivo = request.FILES.get('archivo_respaldo')
    if not archivo:
        messages.error(request, 'Debe seleccionar un archivo de respaldo.')
        return redirect('backup')

    if not archivo.name.endswith('.zip'):
        messages.error(request, 'El archivo debe ser un ZIP.')
        return redirect('backup')

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        zip_path = tmp / 'uploaded.zip'

        with open(str(zip_path), 'wb') as f:
            for chunk in archivo.chunks():
                f.write(chunk)

        try:
            with zipfile.ZipFile(str(zip_path), 'r') as zf:
                zf.extractall(str(tmp))
        except zipfile.BadZipFile:
            messages.error(request, 'El archivo ZIP está corrupto o no es válido.')
            return redirect('backup')

        # Determinar formato del respaldo: nuevo (backup.json) o antiguo (db.sqlite3)
        es_formato_antiguo = (tmp / 'db.sqlite3').exists()
        es_formato_nuevo = (tmp / 'backup.json').exists()

        if not es_formato_antiguo and not es_formato_nuevo:
            messages.error(request, 'El respaldo no contiene una base de datos válida. Debe incluir backup.json o db.sqlite3.')
            return redirect('backup')

        # Si es formato antiguo (SQLite), verificar que el motor actual también sea SQLite
        if es_formato_antiguo and 'postgresql' in settings.DATABASES['default']['ENGINE']:
            messages.error(request, 'El respaldo es de SQLite pero la base de datos actual es PostgreSQL. No se puede restaurar.')
            return redirect('backup')

        # 1. Restaurar la base de datos
        try:
            if es_formato_nuevo:
                # Formato nuevo: limpiar BD y cargar fixture JSON (funciona con PostgreSQL y SQLite)
                call_command('flush', '--noinput', verbosity=0)
                call_command('loaddata', str(tmp / 'backup.json'), verbosity=0)
            elif es_formato_antiguo:
                # Formato antiguo: copiar archivo SQLite directamente (solo SQLite)
                db_destino = Path(settings.DATABASES['default']['NAME'])
                shutil.copy2(str(tmp / 'db.sqlite3'), str(db_destino))
        except Exception as e:
            messages.error(request, f'Error al restaurar la base de datos: {e}')
            return redirect('backup')

        # 3. Restaurar archivos media
        if (tmp / 'media').exists():
            try:
                if settings.MEDIA_ROOT.exists():
                    shutil.rmtree(str(settings.MEDIA_ROOT))
                shutil.copytree(str(tmp / 'media'), str(settings.MEDIA_ROOT))
            except Exception as e:
                messages.error(request, f'Error al restaurar archivos: {e}')
                return redirect('backup')

        timestamp = datetime.now().strftime('%d/%m/%Y %H:%M')
        auditar(request, "RESTAURAR", "Sistema", None,
                f"Restauración desde {archivo.name}", f"Realizada el {timestamp}")

        messages.success(request, 'Respaldo restaurado exitosamente. La base de datos y los archivos han sido recuperados.')
        return redirect('backup')


# ============================================================
# AUDIT LOG VIEWER (Admin only) — Día / Semana / Mes / Año
# ============================================================

def _filtrar_audit_logs(periodo, anio, mes, semana, modelo, accion):
    """Aplica filtros de periodo, modelo y acción a AuditLog y retorna el QuerySet ordenado."""
    qs = AuditLog.objects.all()
    hoy = date.today()

    if modelo:
        qs = qs.filter(modelo__icontains=modelo)
    if accion:
        qs = qs.filter(accion=accion)

    if not anio:
        anio = str(hoy.year)
    if not mes:
        mes = str(hoy.month)

    if periodo == 'dia':
        qs = qs.filter(fecha__date=hoy)
    elif periodo == 'semana' and semana:
        try:
            semana_int = int(semana)
            primer_dia = date(int(anio), 1, 1)
            dias_restar = primer_dia.weekday()
            if dias_restar <= 3:
                lunes_sem1 = primer_dia - timedelta(days=dias_restar)
            else:
                lunes_sem1 = primer_dia + timedelta(days=(7 - dias_restar))
            lunes = lunes_sem1 + timedelta(weeks=semana_int - 1)
            domingo = lunes + timedelta(days=6)
            qs = qs.filter(fecha__date__gte=lunes, fecha__date__lte=domingo)
        except (ValueError, TypeError):
            pass
    elif periodo == 'mes' and mes:
        qs = qs.filter(fecha__year=int(anio), fecha__month=int(mes))
    elif periodo == 'anio' and anio:
        qs = qs.filter(fecha__year=int(anio))

    return qs.order_by("-fecha"), hoy, anio, mes


@permiso_required(perms.AUDITORIA_VER)
def audit_log_list(request):
    """Lista de registros de auditoría con filtros por periodo (día/semana/mes/año), modelo y acción."""
    from django.core.paginator import Paginator

    periodo = request.GET.get('periodo', 'dia')
    anio = request.GET.get('anio', '')
    mes = request.GET.get('mes', '')
    semana = request.GET.get('semana', '')
    modelo = request.GET.get("modelo", "")
    accion = request.GET.get("accion", "")

    qs, hoy, anio, mes = _filtrar_audit_logs(periodo, anio, mes, semana, modelo, accion)

    paginator = Paginator(qs, 25)
    page = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page)
    except Exception:
        page_obj = paginator.page(1)

    # Años disponibles
    anios_disponibles = AuditLog.objects.dates('fecha', 'year', order='DESC')
    if not anios_disponibles:
        anios_disponibles = [hoy]

    # Meses
    meses_opciones = []
    for i in range(1, 13):
        meses_opciones.append((i, calendar.month_name[i].capitalize()))

    context = {
        "logs": page_obj,
        "page_obj": page_obj,
        "paginator": paginator,
        "is_paginated": paginator.num_pages > 1,
        "total_logs": AuditLog.objects.count(),
        "modelo_filtro": modelo,
        "accion_filtro": accion,
        "acciones": AuditLog.ACCIONES,
        "periodo": periodo,
        "anio_actual": anio,
        "mes_actual": mes,
        "semana_actual": semana,
        "anios_disponibles": anios_disponibles,
        "meses_opciones": meses_opciones,
        "hoy": hoy,
    }
    return render(request, "direccion/audit_log_list.html", context)


@permiso_required(perms.AUDITORIA_EXPORTAR_EXCEL)
def exportar_auditoria_excel(request):
    """Exporta los registros de auditoría filtrados a Excel."""
    periodo = request.GET.get('periodo', 'dia')
    anio = request.GET.get('anio', '')
    mes = request.GET.get('mes', '')
    semana = request.GET.get('semana', '')
    modelo = request.GET.get("modelo", "")
    accion = request.GET.get("accion", "")

    qs, _, _, _ = _filtrar_audit_logs(periodo, anio, mes, semana, modelo, accion)

    headers = ["Fecha", "Usuario", "Acción", "Modelo", "Objeto", "Detalle", "IP"]
    rows = []
    for log in qs:
        rows.append([
            log.fecha.strftime("%d/%m/%Y %H:%M"),
            log.username,
            log.get_accion_display(),
            log.modelo,
            log.objeto_repr or "",
            log.detalle or "",
            log.direccion_ip or "",
        ])

    auditar(request, "EXPORTAR", "AuditLog", None, "Exportación Excel",
            f"{qs.count()} registros")
    return _build_excel_response(
        "Auditoría", headers, rows,
        [20, 25, 18, 25, 40, 40, 18], "auditoria.xlsx"
    )


@permiso_required(perms.AUDITORIA_EXPORTAR_PDF)
def exportar_auditoria_pdf(request):
    """Exporta los registros de auditoría filtrados a PDF."""
    periodo = request.GET.get('periodo', 'dia')
    anio = request.GET.get('anio', '')
    mes = request.GET.get('mes', '')
    semana = request.GET.get('semana', '')
    modelo = request.GET.get("modelo", "")
    accion = request.GET.get("accion", "")

    qs, _, _, _ = _filtrar_audit_logs(periodo, anio, mes, semana, modelo, accion)

    headers = ["Fecha", "Usuario", "Acción", "Modelo", "Objeto", "Detalle", "IP"]
    rows = []
    for log in qs:
        rows.append([
            log.fecha.strftime("%d/%m/%Y %H:%M"),
            log.username,
            log.get_accion_display(),
            log.modelo,
            log.objeto_repr or "",
            log.detalle or "",
            log.direccion_ip or "",
        ])

    auditar(request, "EXPORTAR", "AuditLog", None, "Exportación PDF",
            f"{qs.count()} registros")
    return _build_pdf_response(
        "Registros de Auditoría - Reporte",
        headers, rows, "auditoria.pdf"
    )
