from django.shortcuts import redirect
from django.contrib import messages
# ============================================================
# Exportación a Excel y PDF
# ============================================================

from io import BytesIO
from datetime import datetime

from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from ..decorators import permiso_required
from .. import permissions as perms
from django.conf import settings

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.colors import HexColor
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image
)

from ..models import Personal, Investigado
from ..decorators import permiso_required
from ..audit import auditar
from .. import permissions as perms


def _build_excel_response(sheet_title, headers, rows, column_widths, filename):
    """Construye una respuesta HTTP con un archivo Excel (.xlsx) con estilo consistente."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    hf = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="003363", end_color="003363", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center")
    tb = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hf, hfill, ha, tb

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val).border = tb

    for col, w in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _build_pdf_response(title, headers, rows, filename):
    """Construye una respuesta HTTP con un archivo PDF con estilo corporativo."""
    # Buscar el logo en STATICFILES_DIRS (directorio fuente) o STATIC_ROOT (collectstatic)
    _logo_candidates = [
        settings.STATICFILES_DIRS[0] / 'images' / 'logo.png',
        settings.STATIC_ROOT / 'images' / 'logo.png',
    ]
    LOGO_PATH = None
    for _p in _logo_candidates:
        if _p.exists():
            LOGO_PATH = str(_p)
            break

    def _make_logo():
        if LOGO_PATH:
            return Image(LOGO_PATH, width=30, height=30)
        # Si no hay logo disponible, retornar celda vacía
        return Paragraph("", getSampleStyleSheet()["Normal"])

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter),
                            leftMargin=0.5 * inch, rightMargin=0.5 * inch,
                            topMargin=0.6 * inch, bottomMargin=0.5 * inch)
    styles = getSampleStyleSheet()

    elements = []

    # --- Header: logo + title ---
    header_cells = [
        [_make_logo(),
         Paragraph(
             f"<b>{title}</b><br/>"
             f"<font size='7.5' color='#6B7280'>Dirección General - Sistema de Gestión</font>",
             ParagraphStyle("HdrBox", parent=styles["Normal"],
                            fontSize=14, textColor=HexColor("#003363"),
                            spaceBefore=0, spaceAfter=0, leading=18)
         )]
    ]
    hdr_table = Table(header_cells, colWidths=[36, landscape(letter)[0] - inch - 36])
    hdr_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),
        ('RIGHTPADDING', (1, 0), (1, 0), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(hdr_table)

    # Separator line
    elements.append(HRFlowable(width="100%", thickness=1.5,
                                color=HexColor("#003363"),
                                spaceAfter=10, spaceBefore=4))

    # --- Data table ---
    table_data = [headers]
    for row in rows:
        table_data.append([str(c) if c is not None else "" for c in row])

    col_widths = [max(len(str(h)) * 9, 60) for h in headers]
    total_w = sum(col_widths)
    page_w = landscape(letter)[0] - inch
    if total_w > page_w:
        col_widths = [w * page_w / total_w for w in col_widths]

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#003363")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.Color(0, 0, 0, 0.2)),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.Color(1, 1, 1, 1), colors.Color(0.95, 0.95, 0.97, 1)]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 12))

    # Footer
    foot_style = ParagraphStyle(
        "Foot", parent=styles["Normal"],
        fontSize=7, textColor=HexColor("#9CA3AF"), alignment=1
    )
    elements.append(Paragraph(
        f"Generado: {datetime.now():%d/%m/%Y %H:%M} | Total: {len(rows)} registros",
        foot_style
    ))

    doc.build(elements)
    pdf = buf.getvalue()
    buf.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _human_size(size_bytes):
    """Convierte bytes a formato legible (KB, MB, GB)."""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024:
            return f'{size_bytes:.1f} {unit}'
        size_bytes /= 1024
    return f'{size_bytes:.1f} TB'


# ============================================================
# EXPORT TO EXCEL
# ============================================================

@login_required
@permiso_required(perms.EXPORTAR_PERSONAL_EXCEL)
def exportar_personal_excel(request):
    try:
        headers = ["Apellidos", "Nombres", "Cedula", "Telefonos", "Fecha Nacimiento", "Fecha Registro"]
        rows = []
        for p in Personal.objects.filter(activo=True).order_by("apellidos"):
            rows.append([
                p.apellidos,
                p.nombres,
                p.cedula,
                p.telefonos or "",
                p.fecha_nacimiento.strftime("%d/%m/%Y") if p.fecha_nacimiento else "",
                p.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
            ])
        auditar(request, "EXPORTAR", "Personal", None, "Exportación Excel",
                f"{Personal.objects.filter(activo=True).count()} registros")
        return _build_excel_response(
            "Personal", headers, rows,
            [30, 30, 18, 30, 18, 20], "personal_direccion.xlsx"
        )
    except Exception as e:
        messages.error(request, f"Error al exportar Personal a Excel: {str(e)}")
        return redirect('personal_list')


@login_required
@permiso_required(perms.EXPORTAR_INVESTIGADOS_EXCEL)
def exportar_investigados_excel(request):
    try:
        headers = ["Apellidos", "Nombres", "Cedula", "RIF", "Partida Nacimiento", "Entrada Investigacion", "Fecha Registro"]
        rows = []
        for i in Investigado.objects.filter(activo=True).order_by("apellidos"):
            rows.append([
                i.apellidos,
                i.nombres,
                i.cedula or "",
                i.rif or "",
                i.partida_nacimiento or "",
                i.entrada_investigacion or "",
                i.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
            ])
        auditar(request, "EXPORTAR", "Investigado", None, "Exportación Excel",
                f"{Investigado.objects.filter(activo=True).count()} registros")
        return _build_excel_response(
            "Investigados", headers, rows,
            [30, 30, 18, 18, 25, 40, 20], "investigados.xlsx"
        )
    except Exception as e:
        messages.error(request, f"Error al exportar Investigados a Excel: {str(e)}")
        return redirect('investigado_list')


# ============================================================
# EXPORT TO PDF
# ============================================================

@login_required
@permiso_required(perms.EXPORTAR_PERSONAL_PDF)
def exportar_personal_pdf(request):
    qs = Personal.objects.filter(activo=True).order_by("apellidos", "nombres")
    headers = ["Apellidos", "Nombres", "Cédula", "Teléfonos", "Fecha Nac.", "Fecha Registro"]
    rows = []
    for p in qs:
        rows.append([
            p.apellidos,
            p.nombres,
            p.cedula,
            p.telefonos or "",
            p.fecha_nacimiento.strftime("%d/%m/%Y") if p.fecha_nacimiento else "",
            p.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
        ])
    try:
        auditar(request, "EXPORTAR", "Personal", None, "Exportación PDF", f"{qs.count()} registros")
        return _build_pdf_response(
            "Personal de la Dirección General - Reporte",
            headers, rows, "personal_direccion.pdf"
        )
    except Exception as e:
        messages.error(request, f"Error al exportar Personal a PDF: {str(e)}")
        return redirect('personal_list')


@login_required
@permiso_required(perms.EXPORTAR_INVESTIGADOS_PDF)
def exportar_investigados_pdf(request):
    try:
        qs = Investigado.objects.filter(activo=True).order_by("apellidos", "nombres")
        headers = ["Apellidos", "Nombres", "Cédula", "RIF", "Partida Nacimiento", "Entrada Investigación", "Fecha Registro"]
        rows = []
        for i in qs:
            rows.append([
                i.apellidos,
                i.nombres,
                i.cedula or "",
                i.rif or "",
                i.partida_nacimiento or "",
                i.entrada_investigacion or "",
                i.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
            ])
        auditar(request, "EXPORTAR", "Investigado", None, "Exportación PDF", f"{qs.count()} registros")
        return _build_pdf_response(
            "Personas a Investigar - Reporte",
            headers, rows, "investigados.pdf"
        )
    except Exception as e:
        messages.error(request, f"Error al exportar Investigados a PDF: {str(e)}")
        return redirect('investigado_list')


